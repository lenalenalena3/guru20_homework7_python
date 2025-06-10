import pathlib
import zipfile
from pypdf import PdfReader
import csv
from io import TextIOWrapper
from openpyxl import load_workbook

BASE_DIR = pathlib.Path('.')
PATH_FILES_DOWNLOAD = BASE_DIR / 'tmp'
PATH_ARCHIVE = BASE_DIR / 'attachment'
PATH_NAME_ZIP = PATH_ARCHIVE / 'zip_file.zip'
NAME_FILE_PDF = 'file_example_PDF.pdf'
NAME_FILE_CSV = 'file_example_CSV.csv'
NAME_FILE_XLSX = 'file_example_XLSX.xlsx'


def test_create_archive():  # создание архива
    if not PATH_ARCHIVE.exists():  # проверяем существует ли папка
        PATH_ARCHIVE.mkdir()  # создаем папку если её нет
    with zipfile.ZipFile(PATH_NAME_ZIP, 'w') as zip_archive:
        for file in PATH_FILES_DOWNLOAD.iterdir():
            if file.is_file():
                zip_archive.write(file, file.name)  # добавляем файл в архив
    # проверяем, что все файлы добавились
    with zipfile.ZipFile(PATH_NAME_ZIP) as zip_archive:
        print(zip_archive.namelist())
        assert zip_archive.namelist().__len__() == 3


def test_read_pdf_zip():  # чтение и проверка содержимого файла pdf из архива
    with zipfile.ZipFile(PATH_NAME_ZIP) as zip_archive:  # открываем архив
        with zip_archive.open(NAME_FILE_PDF) as pdf_file:  # открываем файл в архиве
            pdf_reader = PdfReader(pdf_file)
            for page in pdf_reader.pages:
                print(page.extract_text())
            number_pages = pdf_reader.get_num_pages() # проверяем количество страниц всего
            assert number_pages == 8
            # проверяем что присутствует определенный текст на первой, второй и последней странице
            assert pdf_reader.pages[0].extract_text().__contains__(
                '2-комн. апарт. \nМосква, район Пресненский, Патриаршие Пруды мкр\nМетро: Пушкинская, 7 мин. пешком\n467 910 000 ₽\n2 700 000 ₽ за м²')
            assert pdf_reader.pages[1].extract_text().__contains__(
                'Архитектурная концепция от бюро "Цимайло, Ляшенко и партнеры" гармонично дополняет шедевры модерна новыми формами и материалами')
            assert pdf_reader.pages[number_pages - 1].extract_text().__contains__(
                '© 2012 - 2025 ЦИАН. Крупнейшая и самая достоверная база данных по аренде и продаже жилой, коммерческой и загородной недвижимости - www.cian.ru')


def test_read_xlsx_zip():  # чтение и проверка содержимого файла xlsx из архива
    with zipfile.ZipFile(PATH_NAME_ZIP) as zip_archive:  # открываем архив
        with zip_archive.open(NAME_FILE_XLSX) as xlsx_file:  # открываем файл в архиве
            workbook = load_workbook(xlsx_file, data_only=True)
            count_sheet = len(workbook.worksheets)
            sheet_names = workbook.sheetnames
            print(sheet_names)
            for sheet_name in sheet_names:
                print(f"*************** Лист: {sheet_name} ************************")
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():  # печатаем содержимое всех ячеек
                    for cell in row:
                        if cell.value != None:
                            print(cell.value)

            assert count_sheet == 2  # проверяем количество листов
            assert sheet_names == ['Sheet1', 'MyLinks']  # проверяем название листов
            for sheet_name in sheet_names:
                # для первого листа проверяем содержимое для строки заголовков, для следующей строки после заголовков и для последней строки
                if sheet_name == 'Sheet1':
                    sheet = workbook[sheet_name]
                    for row_index, row_sheet in enumerate(sheet.iter_rows(values_only=True), start=1):
                        if row_index == 1:
                            assert row_sheet == ('ID', 'Months', None, 'ListMonths', None, None, None)
                        elif row_index == 2:
                            assert row_sheet == (1, 'January', None, 'January', None, None, 'Select Month')
                        elif row_index == 13:
                            assert row_sheet == (9, 'December', None, None, None, None, None)
                # для второго листа проверяем содержимое для первой и последней строк
                if sheet_name == 'MyLinks':
                    sheet = workbook[sheet_name]
                    for row_index, row_sheet in enumerate(sheet.iter_rows(values_only=True), start=1):
                        if row_index == 2:
                            assert row_sheet == (None, 'Contextures Products', None)
                        elif row_index == 17:
                            assert row_sheet == (None, 'Contextures Recommends',
                                                 'Other Excel tools and training, recommended by Debra')


def test_read_csv_zip():  # чтение и проверка содержимого файла csv из архива
    row = {}
    with zipfile.ZipFile(PATH_NAME_ZIP) as zip_archive:  # открываем архив
        with zip_archive.open(NAME_FILE_CSV) as csv_file:  # открываем файл в архиве
            csv_reader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8-sig')))
            count_line = csv_reader.__len__()
            for i in range(count_line):
                row[i] = csv_reader[i]
                print(row[i])
            assert count_line == 20 # проверяем количество строк всего
            # проверяем содержимое заголовков, следующей строки после заголовков и содержимое последней строки
            assert row[0] == ['name', 'phoneNumber', 'email', 'address', 'userAgent', 'hexcolor']
            assert row[1] == ['Lizzie Stanton Sr.', '(494) 333-0427', 'altenwerth.damien@reichert.net',
                              '5577 Jaren Junction Apt. 952\nParisside, WI 27442',
                              'Mozilla/5.0 (X11; Linux x86_64; rv:7.0) Gecko/20100815 Firefox/36.0', '#a45c57']
            assert row[count_line - 1] == ['Mr. Jaiden Johns', '659-533-8311', 'clarkin@stroman.com',
                                      '235 Cormier Union Suite 876\nErnserburgh, SC 61471-3406',
                                      'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/5331 (KHTML, like Gecko) Chrome/36.0.841.0 Mobile Safari/5331',
                                      '#374582']
